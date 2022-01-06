import openpyxl as xl
import csv


file = 'Employee_data.xlsx'

# Load in the workbook
wb = xl.load_workbook(file)

sheet = wb['Sheet1']

old_domaine = 'helpinghands.cm'
new_domaine = 'handsinhand.org'

for i in range (2, sheet.max_row + 1):
      cell = sheet.cell(i, 2) #choosing the column i'm going to work with
      if old_domaine in cell.value : 
        Updated_email = (cell.value).replace(old_domaine, new_domaine) #replacing the old domaine by the new domaine
  
        sheet.cell(i, 2).value = Updated_email#
        
wb.save('Updated_cell_Employee_data.xlsx')

#working now with the csv file
op = open("Employee_data.csv", 'r')
dt = csv.DictReader(op)
print(dt)
up_dt = []
for r in dt:
    print(r)
    row = {'employee name' : r['employee name'],
           'email address' : r['email address'].replace('helpinghands.cm','handsinhand.org'),
           'phone number' : r['phone number'],}
    up_dt.append(row)
print(up_dt)
op.close()

op = open("NewEmployee_data.csv", "w", newline ='')
headers = ['employee name', 'email address', 'phone number']
data = csv.DictWriter(op, delimiter=',', fieldnames= headers)
data.writerow(dict((heads, heads) for heads in headers))
data.writerows(up_dt)
op.close()